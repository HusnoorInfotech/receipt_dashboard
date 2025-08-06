from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
# Create your views here.
from .models import MoneyReceipt
from django.db.models import Sum, Count
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from .models import MoneyReceipt
from num2words import num2words
from django.views.decorators.csrf import csrf_exempt
import urllib.parse


def export_receipts_to_excel(request):
    # Create a new Excel workbook
    wb = Workbook()
    
    # Create sheets for each section
    sections = ['preprimary', 'primary', 'secondary', 'college']
    section_names = {'preprimary': 'Pre-Primary', 'primary': 'Primary', 'secondary': 'Secondary' , 'college':'College'}
    
    # Create a sheet for each section
    for section in sections:
        ws = wb.create_sheet(title=section_names[section])
        ws.append(['Receipt No', 'Student Name', 'Standard', 'Date', 'Total Fees'])
        
        receipts = MoneyReceipt.objects.filter(section=section).order_by('date')
        for receipt in receipts:
            ws.append([
                str(receipt.receipt_no),
                receipt.student_name.capitalize(),
                receipt.standard,
                receipt.date.strftime('%Y-%m-%d'),
                float(receipt.total_fees)
            ])
    
    # Create sheets for each section with computers only
    for section in sections:
        ws = wb.create_sheet(title=f"{section_names[section]} With Computers")
        ws.append(['Receipt No', 'Student Name', 'Standard', 'Date', 'Total Fees'])
        
        receipts = MoneyReceipt.objects.filter(section=section, with_computers=True).order_by('date')
        for receipt in receipts:
            ws.append([
                str(receipt.receipt_no),
                receipt.student_name,
                receipt.standard,
                receipt.date.strftime('%Y-%m-%d'),
                float(receipt.total_fees)
            ])
    
    # Remove the default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create HTTP response with Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=money_receipts.xlsx'
    wb.save(response)
    
    return response







def index(request):
    total_receipt = MoneyReceipt.objects.count()
    total_paid = MoneyReceipt.objects.aggregate(total=Sum('total_fees'))
    receipt_data = MoneyReceipt.objects.order_by('date')[::-1][:4]
    # print(receipt_data)s
    return render(request, 'index.html',{
        'total_receipt':total_receipt,
        'total_paid':total_paid['total'],
        'receipt_data':receipt_data,
        })
    # return HttpResponse(all_data)


def showform(request):
    return render(request,'forms.html')


def submitform(request):
    if request.method == 'POST':
        student_name = request.POST.get('studentName')
        student_section = request.POST.get('section')
        student_standard = request.POST.get('standard')
        student_fee = request.POST.get('feeAmount')
        student_computer = request.POST.get('computerOption')

        try:

            # saving data in database
            saveData = MoneyReceipt.objects.create(
                section=student_section,
                with_computers = True if student_computer == 'on' else False,
                student_name = student_name,
                standard = student_standard,
                total_fees = student_fee
                )
            
            
            messages.success(request,"Formsubmitted successfully")
            return redirect('index')
        except Exception as err:
            messages.error(request,"data not saved")
            return redirect('index')

        
    else:
        return render(request,'forms.html')


def printform(request,receipt_no):
    try:
        receipt = MoneyReceipt.objects.get(receipt_no=receipt_no)
        feesInWords = num2words(receipt.total_fees,lang='en_IN')
        return render(request,'printform.html',{
            "receipt":receipt,
            "feesinwords":feesInWords
            })
    except MoneyReceipt.DoesNotExist:
        error_message = f"MoneyReceipt with receipt_no '{receipt_no}' does not exist."
        # print(error_message)
        return render(request, 'printform.html', {"error": error_message})
    except Exception as err:
        # print(err)
        return render(request,'printform.html',{"error":"An unexpected error occurred."})


# update the form

def showformtoedit(request,receipt_no):
    try:
        receipt = MoneyReceipt.objects.get(receipt_no=receipt_no)
        # print(receipt)
        return render(request,'forms.html',{"receipt":receipt})
    except MoneyReceipt.DoesNotExist:
        error_message = f"MoneyReceipt with receipt_no '{receipt_no}' does not exist."
        return render(request, 'printform.html', {"error": error_message})
    except Exception as err:
        return render(request, 'printform.html',{"error":err})


# @csrf_exempt  # Use with caution; consider using proper CSRF protection
def updateform(request,receipt_no):
    try:
        receipt = MoneyReceipt.objects.get(receipt_no=receipt_no)

        if request.method == 'POST':
            student_name = request.POST.get('studentName')
            # print(student_name)
            student_section = request.POST.get('section')
            student_standard = request.POST.get('standard')
            student_fee = request.POST.get('feeAmount')
            student_computer = request.POST.get('computerOption')
            if request.POST.get('_method') == 'PUT':
                # Update the receipt fields
                receipt.section = student_section
                receipt.with_computers = True if student_computer == 'on' else False
                # receipt.student_name = student_name
                receipt.standard = student_standard
                receipt.total_fees = student_fee
                # Save the updated receipt
                receipt.save()
                
                messages.success(request,"data change")
                return redirect('index')
        else:
            return redirect('index')
    except MoneyReceipt.DoesNotExist:
        messages.error(request, "Receipt not found")
        return redirect('index')
    except Exception as err:
        messages.error(request,err)
        return redirect('index')


